"""
Export functionality for Maynord Calculator
Creates professional reports of calculations
"""

import io
import os
import tempfile
from datetime import datetime
from typing import Optional, Dict, Any

from core.maynord import MaynordResult
from core.gradation import get_complete_gradation_summary


def generate_clean_chart_for_pdf(result: MaynordResult, filepath: str):
    """
    Generate a clean chart specifically for PDF export
    with properly positioned annotations that don't overlap
    """
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend for export
    import matplotlib.pyplot as plt
    import numpy as np

    # Create figure with white background
    fig, ax = plt.subplots(figsize=(8, 5), dpi=150)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#f8fafc')

    # Calculate gradation points
    d15 = result.d30 * 0.70
    d30 = result.d30
    d50 = result.d50
    d85 = result.d30 * 1.70
    d100 = result.d100

    diameters = [d15, d30, d50, d85, d100]
    percentages = [15, 30, 50, 85, 100]

    # Set up axes
    ax.set_xscale('log')
    x_min = max(5, d15 * 0.4)
    x_max = d100 * 2.0
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(0, 110)

    ax.set_xlabel('Diametre (mm)', fontsize=11, fontweight='bold', color='#334155')
    ax.set_ylabel('Passant cumule (%)', fontsize=11, fontweight='bold', color='#334155')
    ax.set_title('Courbe granulometrique', fontsize=14, fontweight='bold', color='#1e40af', pad=15)

    # Grid
    ax.grid(True, which='major', linestyle='-', alpha=0.3, color='#cbd5e1')
    ax.grid(True, which='minor', linestyle=':', alpha=0.2, color='#e2e8f0')
    ax.tick_params(colors='#334155', labelsize=10)

    # Plot curve
    ax.plot(diameters, percentages, 'o-', color='#2563eb',
            linewidth=2.5, markersize=10, markerfacecolor='white',
            markeredgewidth=2.5, label='Gradation calculee', zorder=5)

    # Fill area
    ax.fill_between(diameters, [0]*5, percentages, alpha=0.1, color='#2563eb')

    # Vertical reference lines
    ax.axvline(x=d30, color='#16a34a', linestyle='--', alpha=0.7, linewidth=1.5)
    ax.axvline(x=d50, color='#f59e0b', linestyle='--', alpha=0.7, linewidth=1.5)
    ax.axvline(x=d100, color='#dc2626', linestyle='--', alpha=0.7, linewidth=1.5)

    # Horizontal reference lines
    ax.axhline(y=30, color='#e2e8f0', linestyle='-', alpha=0.5, linewidth=0.5)
    ax.axhline(y=50, color='#e2e8f0', linestyle='-', alpha=0.5, linewidth=0.5)
    ax.axhline(y=100, color='#e2e8f0', linestyle='-', alpha=0.5, linewidth=0.5)

    # Simple text annotations (positioned to avoid overlap)
    # D30 annotation - left side
    ax.annotate(f'D30\n{d30:.0f} mm',
                xy=(d30, 30),
                xytext=(d30 * 0.35, 42),
                fontsize=10, fontweight='bold', color='#16a34a',
                ha='center', va='bottom',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                         edgecolor='#16a34a', alpha=0.95),
                arrowprops=dict(arrowstyle='->', color='#16a34a', lw=1.5))

    # D50 annotation - right side upper
    ax.annotate(f'D50\n{d50:.0f} mm',
                xy=(d50, 50),
                xytext=(d50 * 2.5, 58),
                fontsize=10, fontweight='bold', color='#f59e0b',
                ha='center', va='bottom',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                         edgecolor='#f59e0b', alpha=0.95),
                arrowprops=dict(arrowstyle='->', color='#f59e0b', lw=1.5))

    # D100 annotation - above
    ax.annotate(f'D100\n{d100:.0f} mm',
                xy=(d100, 100),
                xytext=(d100 * 0.5, 102),
                fontsize=10, fontweight='bold', color='#dc2626',
                ha='center', va='bottom',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white',
                         edgecolor='#dc2626', alpha=0.95),
                arrowprops=dict(arrowstyle='->', color='#dc2626', lw=1.5))

    # Legend
    ax.legend(loc='lower right', fontsize=10, framealpha=0.95,
             facecolor='white', edgecolor='#e2e8f0')

    # Tight layout
    plt.tight_layout()

    # Save
    fig.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
    plt.close(fig)

    return filepath


def export_calculation_to_excel(
    result: MaynordResult,
    input_params: Dict[str, Any],
    filepath: str
):
    """
    Export a single calculation to Excel with professional formatting

    Args:
        result: MaynordResult from calculation
        input_params: Dict with velocity, depth, config info
        filepath: Output file path
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Calcul"

    # Styles
    title_font = Font(bold=True, size=16, color="1E40AF")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    section_font = Font(bold=True, size=12, color="1E40AF")
    value_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    row = 1

    # Title
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="RAPPORT DE CALCUL - DIMENSIONNEMENT ENROCHEMENTS").font = title_font
    row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value=f"Methode Maynord (USACE) - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    row += 2

    # Input Parameters Section
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="PARAMETRES D'ENTREE").font = section_font
    row += 1

    params = [
        ("Vitesse V", f"{input_params.get('velocity', 0):.2f} m/s"),
        ("Profondeur D", f"{input_params.get('depth', 0):.2f} m"),
        ("Type de section", input_params.get('section_type', 'Lit')),
        ("Configuration chenal", input_params.get('channel_type', 'Droit')),
    ]
    if input_params.get('slope'):
        params.append(("Pente talus", input_params.get('slope')))
    if input_params.get('rw_ratio'):
        params.append(("Ratio R/W", f"{input_params.get('rw_ratio'):.1f}"))

    for label, value in params:
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=value).font = value_font
        ws.cell(row=row, column=2).border = border
        row += 1

    row += 1

    # Coefficients Section
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="COEFFICIENTS").font = section_font
    row += 1

    coeffs = result.coefficients
    coeff_data = [
        ("SF (Facteur securite)", f"{coeffs.get('SF', 1.1):.2f}"),
        ("Cs (Stabilite)", f"{coeffs.get('Cs', 0.375):.3f}"),
        ("Cv (Distribution vitesse)", f"{coeffs.get('Cv', 1.0):.3f}"),
        ("K1 (Correction pente)", f"{coeffs.get('K1', 1.0):.3f}"),
        ("Ss (Densite roche)", f"{coeffs.get('Ss', 2.65):.2f}"),
        ("CT (Epaisseur)", f"{coeffs.get('CT', 1.0):.2f}"),
    ]

    for label, value in coeff_data:
        ws.cell(row=row, column=1, value=label).border = border
        ws.cell(row=row, column=2, value=value).font = value_font
        ws.cell(row=row, column=2).border = border
        row += 1

    row += 1

    # Results Section
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="RESULTATS").font = section_font
    row += 1

    # Headers
    for col, header in enumerate(["Parametre", "Valeur", "Unite"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Get gradation summary
    summary = get_complete_gradation_summary(result.d30, coeffs.get('Ss', 2.65))

    results_data = [
        ("D30 (calcule)", f"{result.d30:.1f}", "mm"),
        ("D50 (estime)", f"{result.d50:.1f}", "mm"),
        ("D100 (estime)", f"{result.d100:.1f}", "mm"),
        ("Masse D30", f"{result.mass_d30:.1f}" if result.mass_d30 < 1000 else f"{result.mass_d30/1000:.2f}", "kg" if result.mass_d30 < 1000 else "t"),
        ("Masse D50", f"{result.mass_d50:.1f}" if result.mass_d50 < 1000 else f"{result.mass_d50/1000:.2f}", "kg" if result.mass_d50 < 1000 else "t"),
        ("Masse D100", f"{result.mass_d100:.1f}" if result.mass_d100 < 1000 else f"{result.mass_d100/1000:.2f}", "kg" if result.mass_d100 < 1000 else "t"),
        ("Epaisseur couche", f"{result.thickness:.1f}", "cm"),
        ("Classe gradation", summary.usace_class, "USACE"),
        ("Nombre de Froude", f"{result.froude_number:.3f}", "-"),
        ("Masse par m2", f"{summary.mass_per_m2:.0f}", "kg/m2"),
    ]

    for label, value, unit in results_data:
        ws.cell(row=row, column=1, value=label).border = border
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = value_font
        cell.border = border
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=unit).border = border
        row += 1

    row += 2

    # Status
    status = "STABLE" if result.froude_number <= 1.2 and result.d100 <= 1500 else "LIMITE"
    ws.merge_cells(f'A{row}:C{row}')
    status_cell = ws.cell(row=row, column=1, value=f"STATUS: {status}")
    status_cell.font = Font(bold=True, size=14, color="16A34A" if status == "STABLE" else "F59E0B")
    row += 2

    # Formula reference
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1, value="Equation: d30 = SF x Cs x Cv x CT x D x [V/sqrt(gD(Ss-1))]^2.5 / K1")
    row += 1
    ws.cell(row=row, column=1, value="Reference: Maynord S.T. (1988) - USACE Technical Report HL-88-4")

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    wb.save(filepath)


def export_calculation_to_pdf(
    result: MaynordResult,
    input_params: Dict[str, Any],
    filepath: str,
    chart_image_path: Optional[str] = None
):
    """
    Export a single calculation to professional PDF report

    Args:
        result: MaynordResult from calculation
        input_params: Dict with velocity, depth, config info
        filepath: Output file path
        chart_image_path: Optional path to chart image
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    # Colors
    primary_color = colors.HexColor('#1e40af')
    success_color = colors.HexColor('#16a34a')
    warning_color = colors.HexColor('#f59e0b')
    bg_light = colors.HexColor('#f1f5f9')

    # Create document
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )

    # Styles
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=primary_color,
        spaceAfter=5,
        alignment=TA_CENTER
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=15,
        alignment=TA_CENTER
    )

    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=primary_color,
        spaceBefore=15,
        spaceAfter=8,
        borderPadding=5
    )

    normal_style = styles['Normal']

    elements = []

    # === HEADER ===
    elements.append(Paragraph("RAPPORT DE CALCUL", title_style))
    elements.append(Paragraph("Dimensionnement d'enrochements - Methode Maynord (USACE)", subtitle_style))
    elements.append(Paragraph(f"Genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=primary_color, spaceAfter=15))

    # === STATUS BANNER ===
    is_stable = result.froude_number <= 1.2 and result.d100 <= 1500
    status_text = "STABLE" if is_stable else "LIMITE"
    status_color = success_color if is_stable else warning_color

    status_table = Table([[status_text]], colWidths=[5*cm])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), status_color),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ROUNDEDCORNERS', [5, 5, 5, 5]),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 15))

    # === TWO COLUMN LAYOUT: Params + Coefficients ===
    # Left: Input params | Right: Coefficients

    params_data = [
        ["PARAMETRES D'ENTREE", ""],
        ["Vitesse V", f"{input_params.get('velocity', 0):.2f} m/s"],
        ["Profondeur D", f"{input_params.get('depth', 0):.2f} m"],
        ["Section", input_params.get('section_type', 'Lit')],
        ["Chenal", input_params.get('channel_type', 'Droit')],
    ]
    if input_params.get('slope'):
        params_data.append(["Pente", input_params.get('slope')])
    if input_params.get('rw_ratio'):
        params_data.append(["R/W", f"{input_params.get('rw_ratio'):.1f}"])

    coeffs = result.coefficients
    coeffs_data = [
        ["COEFFICIENTS", ""],
        ["SF", f"{coeffs.get('SF', 1.1):.2f}"],
        ["Cs", f"{coeffs.get('Cs', 0.375):.3f}"],
        ["Cv", f"{coeffs.get('Cv', 1.0):.3f}"],
        ["K1", f"{coeffs.get('K1', 1.0):.3f}"],
        ["Ss", f"{coeffs.get('Ss', 2.65):.2f}"],
        ["CT", f"{coeffs.get('CT', 1.0):.2f}"],
    ]

    # Create side-by-side tables
    params_table = Table(params_data, colWidths=[4*cm, 3*cm])
    params_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('GRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, bg_light]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    coeffs_table = Table(coeffs_data, colWidths=[4*cm, 3*cm])
    coeffs_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('GRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, bg_light]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    # Combine in outer table
    outer_table = Table([[params_table, coeffs_table]], colWidths=[8.5*cm, 8.5*cm])
    outer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(outer_table)
    elements.append(Spacer(1, 15))

    # === RESULTS TABLE ===
    summary = get_complete_gradation_summary(result.d30, coeffs.get('Ss', 2.65))

    def format_mass(m):
        """Returns (value, unit) tuple"""
        if m < 1000:
            return f"{m:.1f}", "kg"
        else:
            return f"{m/1000:.2f}", "t"

    m_d30_val, m_d30_unit = format_mass(result.mass_d30)
    m_d50_val, m_d50_unit = format_mass(result.mass_d50)
    m_d100_val, m_d100_unit = format_mass(result.mass_d100)

    results_data = [
        ["RESULTATS DE DIMENSIONNEMENT", "", ""],
        ["Diametre D30 (calcule)", f"{result.d30:.1f}", "mm"],
        ["Diametre D50 (estime)", f"{result.d50:.1f}", "mm"],
        ["Diametre D100 (estime)", f"{result.d100:.1f}", "mm"],
        ["", "", ""],
        ["Masse unitaire D30", m_d30_val, m_d30_unit],
        ["Masse unitaire D50", m_d50_val, m_d50_unit],
        ["Masse unitaire D100", m_d100_val, m_d100_unit],
        ["", "", ""],
        ["Epaisseur de couche", f"{result.thickness:.1f}", "cm"],
        ["Classe de gradation", summary.usace_class, "USACE"],
        ["Nombre de Froude", f"{result.froude_number:.3f}", ""],
        ["Masse par m2", f"{summary.mass_per_m2:.0f}", "kg/m2"],
    ]

    results_table = Table(results_data, colWidths=[8*cm, 4*cm, 3*cm])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('SPAN', (0, 0), (-1, 0)),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('GRID', (0, 1), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, bg_light]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        # Empty rows styling
        ('BACKGROUND', (0, 4), (-1, 4), colors.white),
        ('BACKGROUND', (0, 8), (-1, 8), colors.white),
        ('LINEABOVE', (0, 4), (-1, 4), 0, colors.white),
        ('LINEBELOW', (0, 4), (-1, 4), 0, colors.white),
    ]))
    elements.append(results_table)
    elements.append(Spacer(1, 15))

    # === CHART IMAGE ===
    # Generate a clean chart specifically for PDF (ignores chart_image_path)
    pdf_chart_path = os.path.join(tempfile.gettempdir(), 'maynord_pdf_chart.png')
    generate_clean_chart_for_pdf(result, pdf_chart_path)

    if os.path.exists(pdf_chart_path):
        elements.append(Paragraph("Courbe granulometrique", section_style))
        img = Image(pdf_chart_path, width=16*cm, height=10*cm)
        elements.append(img)
        elements.append(Spacer(1, 10))

    # === FORMULA REFERENCE ===
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0'), spaceBefore=10, spaceAfter=10))

    formula_style = ParagraphStyle(
        'Formula',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey
    )

    elements.append(Paragraph(
        "<b>Equation de Maynord:</b> d30 = SF x Cs x Cv x CT x D x [V / sqrt(g x D x (Ss-1))]^2.5 / K1",
        formula_style
    ))
    elements.append(Paragraph(
        "<b>Reference:</b> Maynord S.T. (1988) - Stable Riprap Size for Open Channel Flows - USACE Technical Report HL-88-4",
        formula_style
    ))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        f"<i>Rapport genere par Maynord Calculator v1.0 - Developpe par Marc Zermatten - GeoMind 2025</i>",
        formula_style
    ))

    # Build PDF
    doc.build(elements)


# Keep old functions for compatibility but redirect to new ones
def export_to_excel(project_manager, filepath: str):
    """Legacy function - exports last calculation if available"""
    project = project_manager.project
    if project.calculations:
        last_calc = project.calculations[-1]
        # Create a minimal result object
        from core.maynord import MaynordResult
        result_dict = last_calc.get('result', {})
        coeffs = last_calc.get('coefficients', {})

        # This is a simplified export - the new function should be used directly
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dernier calcul"
        ws.cell(row=1, column=1, value="Voir export_calculation_to_excel pour rapport complet")
        wb.save(filepath)


def export_to_pdf(project_manager, filepath: str):
    """Legacy function - exports last calculation if available"""
    project = project_manager.project
    if project.calculations:
        # Simplified fallback
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [Paragraph("Voir export_calculation_to_pdf pour rapport complet", styles['Normal'])]
        doc.build(elements)
